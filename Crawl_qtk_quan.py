#! /usr/bin/env python
#-*-coding:utf-8-*-

import time
import urllib2, os.path as osp
import urllib
import sys  
import re   
import requests
import HTMLParser
from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from time import clock


headers = {'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'}
def GetHtml(url):
   r=requests.get(url,headers=headers)
   text=r.content
   return text
def GetUrlAndQuan(html):
    #print html
    soup = BeautifulSoup(html,'html.parser')
    h    = soup.find_all('body')
    soup = BeautifulSoup(str(h[0]),'html.parser')
    h    = soup.find_all('div',attrs={"class": "goods-list"})
    result=[]
    for each_div in h:
       soup = BeautifulSoup(str(each_div),'html.parser')
       items= soup.find_all('li',attrs={"class": re.compile("fadein")})
       #items= soup.select('li[class]')
       for item in items:
          try:
             result.append(get_items_and_quan(str(item)))
          except:
             print "error"
    return result
def get_items_and_quan(item):
   item_url=re.findall("(?isu)(amp;id\=\d+)",item)
   item_url=list(set(item_url))
   quan_url=re.findall("(?isu)(https\://taoquan.taobao.com/coupon/unify\_apply.htm\?sellerId=\d+\&[0-9a-zA-Z]+)",item)
   quan_url=list(set(quan_url))
   quan_ids =re.findall("(?isu)(activityId\=[0-9a-zA-Z]+)",item)
   quan_ids=list(set(quan_ids))
   quan_urls=[]
   for quan_id in quan_ids:
      quan=quan_url[0]+'&'+quan_id
      quan_urls.append(quan)
   soup  = BeautifulSoup(item,'html.parser')
   items_and_quan=[]
   string= soup.find_all('span',attrs={"class": "type"})
   items_and_quan.append(string[0].string)
   items_and_quan.append( item_url)
   items_and_quan.append( quan_urls)
   return items_and_quan
#url ='http://www.qingtaoke.com/?r=coupon/index'
url ='http://www.qingtaoke.com/quan?r=coupon%2Findex&page='
start=clock()
wb = Workbook()
ws = wb.worksheets[0]
ws.title = u"宝贝数据"
i=1
for k in range(1,5):
   url='http://www.qingtaoke.com/quan?r=coupon%2Findex&page='+str(k)
   result=GetUrlAndQuan(GetHtml(url))
   for data in result:
      try:
         ws.cell(row=i, column=1).value=data[0]
         ws.cell(row=i, column=2).value=(data[1][0].split('='))[-1]
         ws.cell(row=i, column=3).value=str(data[2])
         for j in range(len(data[2])):
            ws.cell(row=i, column=3+j).value=str(data[2][j])
         i=i+1
      except:
         print "error data"
FileName =time.strftime('%Y-%m-%d',time.localtime(time.time()))+'quan.xlsx'
#file_name = 'test323.xlsx'
wb.save(FileName)
endt=clock()
print "the time this code uses is %lf!!!!" %(endt-start)
